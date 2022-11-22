#%%
from cProfile import label
from netCDF4 import Dataset,num2date 
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timezone
import time
import xarray as xr
import utm
import openpyxl
import pandas as pd
import json
import re
from itertools import islice
from mpl_toolkits.basemap import Basemap
from mpl_toolkits.axes_grid1.inset_locator import zoomed_inset_axes
from mpl_toolkits.axes_grid1.inset_locator import mark_inset
from matplotlib import patches
#%%
def dms2dd(degrees, minutes, seconds, direction='E'):
    dd = degrees + minutes/60 + seconds/(60*60)
    if direction == 'S' or direction == 'W':
        dd *= -1
    return dd

def dd2dms(deg):
    d = int(deg)
    md = abs(deg - d) * 60
    m = int(md)
    sd = (md - m) * 60
    return [d, m, sd]

def parse_dms(dms):
    parts = re.split('[^\d\w]+', dms)
    lat = dms2dd(parts[0], parts[1], parts[2], parts[3])
    lng = dms2dd(parts[4], parts[5], parts[6], parts[7])

    return (lat, lng)
# %%

# %%

def create_well(str):
    tmp=re.split(', |:',str)
    X=float(tmp[1])
    Y=float(tmp[3])
    latlon=utm.to_latlon(X, Y, 32,'V')
    out=xr.Dataset(data_vars={
        "X":(float(X)),
        "Y":(float(Y)),
        "lat":(latlon[0]),
        "lon":(latlon[1])}
    )
    return out

# %%
def read_NPD_xlsx_raw(file='Export.xlsx'):
    wb = openpyxl.load_workbook(filename = file)
    try:
        ws = wb['Wellbores, all']
    except:
        ws=wb.worksheets[0]
    
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df.set_index(df['OBJECTID'],inplace=True)
    cols=['NS UTM [m]','EW UTM [m]','NS degrees','NS minutes','NS seconds','EW degrees','EW minutes','EW seconds','NS decimal degrees','EW decimal degrees','UTM zone','EW code','NS code','Water depth [m]','Well name']
    wells=df[cols].to_xarray()

    return wells

def read_NPD_xlsx_old(file='Export.xlsx', method='index'):
    wb = openpyxl.load_workbook(filename = file)
    ws = wb['Wellbores, all']
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    wells=[]
    for id in df.index:
        wells.append(create_well(id))
    wells=xr.concat(wells, dim='wells')
    return wells
#%%
def read_NPD_xlsx(file='Export.xlsx'):
    wb = openpyxl.load_workbook(filename = file)
    ws = wb['Wellbores, all']
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df.set_index(df['OBJECTID'],inplace=True)
    cols=['NS UTM [m]','EW UTM [m]','NS degrees','NS minutes','NS seconds','EW degrees','EW minutes','EW seconds','NS decimal degrees','EW decimal degrees','UTM zone','EW code','NS code','Water depth [m]','Well name']
    wells=df[cols].to_xarray()
    wells['lat']=(('OBJECTID'),dms2dd(wells['NS degrees'].values, wells['NS minutes'].values, wells['NS seconds'].values))
    wells['lon']=(('OBJECTID'),dms2dd(wells['EW degrees'].values, wells['EW minutes'].values, wells['EW seconds'].values, 'E'))
    X,Y,_,_=utm.from_latlon(wells['lon'].values,wells['lat'].values)
    wells['X']=(('OBJECTID'),X)
    wells['Y']=(('OBJECTID'),Y)
    return wells

def NPD_xlsx_to_panda(path='/home/guttorm/Github/ACTOM/ACTOM_sites/North_Sea/Norwegian site/', 
    file='wellbore_coordinates.xlsx'):
    filenm='file:' + path + file
    data=pd.read_excel(path+file)#, usecols=cols)
    return data
# %%

# %%
troll_lat=60.645556
troll_lon=3.726389


# %% Read coordinates from Sarah
def read_sarah(file='leak_coordinates.txt'):
    E=[]
    N=[]
    with open(file) as lines:
        for line in lines:
            e,w=  [float(x) for x in line.split()]
            E.append(e)
            N.append(w)

    
    
    lat, lon= utm.to_latlon(np.array(E),np.array(N),31,'V')   
    data=xr.Dataset(
        data_vars=dict(
            X=('num',E),
            Y=('num',N),
            lat=('num',lat),
            lon=('num',lon))) 
    return data


# set up orthographic map projection with
# perspective of satellite looking down at 50N, 100W.
# use low resolution coastlines.

def make_map_NS(wells,Nother=None, topo=None,savefig=False,
    llcrnrlon=3.,
    llcrnrlat=60.5,
    urcrnrlon=4,
    urcrnrlat=60.9):

    lllon = 1.0
    urlon = 6.7
    urlat = 62.5    
    lllat = 57.9
    fig = plt.figure()
    ax = fig.add_subplot(111)

    map = Basemap(
        projection='cyl',
        lat_0=60,lon_0=4,
        resolution='h',
        llcrnrlat=lllat, 
        urcrnrlat=urlat, 
        llcrnrlon=lllon, 
        urcrnrlon=urlon
        )
# draw coastlines, country boundaries, fill continents.
    map.etopo(scale=0.9,alpha=0.5)
    map.drawcoastlines(linewidth=0.25, color='#cd5b45')
    map.drawcountries(linewidth=0.25)
    map.fillcontinents(color='coral',lake_color='aqua')


#xmin, ymin = map(lllon, lllat)
#xmax, ymax = map(urlon, urlat)
    dotsize=0.1
    alpha=0.5
    lons=wells['EW decimal degrees'].values
    lats=wells['NS decimal degrees'].values
    x, y = map(lons, lats)  # transform coordinates
    plt.scatter(x, y, dotsize, marker='o', alpha=alpha,color='Blue') 

    

    x, y = map(5.32415, 60.39299) # Bergen
    plt.scatter(x, y, 20, marker='o', color='Black')
    plt.annotate('Bergen', xy=(x, y))

# Haugesund 59.41378, 5.268
    x, y = map(5.268, 59.41378) # Haugesund
    plt.scatter(x, y, 20, marker='o', color='Black')
    plt.annotate('Haugesund', xy=(x, y))#, weight='bold')


#Ålesund 62.47225, 6.15492
#x, y = map(6.15492, 62.47225) # Ålesund
#plt.scatter(x, y, 20, marker='o', color='Black')
#plt.annotate('Ålesund', xy=(x, y))

#Mongstad 60.81129 5.032976
    x, y = map(5.032976, 60.81129) # Mongstad
    plt.scatter(x, y, 20, marker='o', color='Black')
    plt.annotate('Mongstad', xy=(x, y))

#Stavanger 58.969975, 5.733107
    x, y = map(5.733107, 58.969975) # Stavanger
    plt.scatter(x, y, 20, marker='o')
    plt.annotate('Stavanger', xy=(x, y), color='Black')
    
    
    width=0.036126320292537706
    height=0.013621245670790927
    angle=0
    x, y = map(2.731416, 58.188087) #Sleipner
    e2 = patches.Ellipse((x, y), width, height,
                     angle=angle, linewidth=2, fill=True, color='Green',  zorder=2)
    ax.add_patch(e2)
    plt.annotate('Sleipner', xy=(x, y), color='Black')

    x, y = map(3.666664, 60.666664 ) #Long Ship
    e3 = patches.Ellipse((x, y), width, height,
                     angle=angle, linewidth=2, fill=True, color='Green',  zorder=2)
    ax.add_patch(e3)
    plt.annotate('Long Ship', xy=(x, y), color='Black')

    for ii in np.arange(Nother):
        x=2+random.uniform(0, 1)*2
        y = 58+random.uniform(0, 1)*4
        e3 = patches.Ellipse((x, y), width, height,
                     angle=angle, linewidth=2, fill=True, color='Green',  zorder=2)
        ax.add_patch(e3)
    plt.show()
   # if savefig:
    #    fig.savefig('NS_map.png')

def read_topo(path='/home/guttorm/Github/ACTOM/ACTOM_sites/North_Sea/Norwegian site/', 
    file='NS_vel201903.nc'):

    data=xr.open_dataset(path + file).isel(time=0).h

    return data

#%%
pos=read_sarah()
wells=NPD_xlsx_to_panda()

#%%
topo=read_topo()
# %%
make_map_NS(wells=wells,Nother=10, topo=topo)
# %%
lon, lat = (4.733107, 58.969975)
x,y,z,q=utm.from_latlon(lon,lat)

xu=x+1500
yu=y+4000

nlon,nlat=utm.to_latlon(xu,yu,z,q)
# %%
